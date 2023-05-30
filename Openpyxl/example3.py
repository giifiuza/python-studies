from openpyxl import load_workbook

wb1 = load_workbook('professores.xlsx')
p1 = wb1['Sheet1']
doutores = []

for i in range(1, 30):
    if p1.cell(row=i+1, column=3).value == "DOUTORADO":
        doutores.append(p1.cell(row=i, column=2).value)

for i in range(len(doutores)):
    x = doutores[i]
    print(x)

wb = load_workbook('teste.xlsx')
p = wb['P3']

for i in range(2, 20):
    p[f'A{i+1}'] = i+1
    for i in range(2, len(doutores)):
        x = doutores[i]
        p[f'B{i+1}'] = x
p["B1"] = "Nome"
wb.save('teste.xlsx')