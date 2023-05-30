from openpyxl import load_workbook

wb = load_workbook('professores.xlsx')
p = wb['Sheet1']
doutores = []

for i in range(1, 30):
    if p.cell(row=i+1, column=3).value == "DOUTORADO":
        doutores.append(p.cell(row=i, column=2).value)

for i in range(len(doutores)):
    print(doutores[i])

